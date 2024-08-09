import telebot
import random
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# جایگزینی 'TOKEN' با توکن ربات خودتان
bot = telebot.TeleBot("Your_API")

# تنظیمات اولیه برای فایل اکسل
workbook = Workbook()
sheet = workbook.active
sheet.title = "Log Data"
headers = ["DateTime", "UserID", "Username", "Message", "Log"]
for idx, header in enumerate(headers, 1):
    sheet[f"{get_column_letter(idx)}1"] = header

# ذخیره اعداد و تلاش‌ها برای هر کاربر
users_data = {}

# تنظیمات لاگ‌گذاری
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("bot.log"),
        logging.StreamHandler()
    ]
)

def log_to_excel(user_id, username, message, log_message):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = [current_time, user_id, username, message, log_message]
    sheet.append(new_row)
    workbook.save("bot_log.xlsx")

@bot.message_handler(commands=['start'])
def send_welcome(message):
    b = random.randint(1, 50)
    users_data[message.chat.id] = {'number': b, 'attempts': 0}
    log_message = "Game started, number generated."
    logging.info(f"{log_message} UserID: {message.chat.id}, Username: {message.from_user.username},javab dorost:{b}")
    log_to_excel(message.chat.id, message.from_user.username, message.text, log_message)
    bot.reply_to(message, "یک عدد بین 1 تا 50 انتخاب شده است. شما 5 فرصت دارید تا آن را حدس بزنید.")

@bot.message_handler(func=lambda message: True)
def guess_number(message):
    chat_id = message.chat.id
    if chat_id not in users_data:
        log_message = "Game not started but user sent a message."
        logging.info(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Message: {message.text}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, "ابتدا با دستور /start بازی را شروع کنید.")
        return

    try:
        a = int(message.text)
    except ValueError:
        log_message = "Invalid input (non-integer) received."
        logging.warning(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Message: {message.text}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, "لطفاً یک عدد معتبر وارد کنید.")
        return

    user_data = users_data[chat_id]
    user_data['attempts'] += 1
    attempts = user_data['attempts']
    b = user_data['number']

    if a < b:
        log_message = "User guessed lower."
        logging.info(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Guess: {a}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, "عدد شما کوچکتر است.")
    elif a > b:
        log_message = "User guessed higher."
        logging.info(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Guess: {a}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, "عدد شما بزرگتر است.")
    else:
        log_message = "User guessed correctly."
        logging.info(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Correct Answer: {b}, Attempts: {attempts}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, f"آفرین! جواب {b} درست بود. تعداد تلاش: {attempts}")
        del users_data[chat_id]
        return

    if attempts == 5:
        log_message = "User used all attempts."
        logging.info(f"{log_message} UserID: {chat_id}, Username: {message.from_user.username}, Correct Answer: {b}")
        log_to_excel(chat_id, message.from_user.username, message.text, log_message)
        bot.reply_to(message, f"متاسفانه فرصت‌های شما تمام شد. عدد مورد نظر: {b}")
        del users_data[chat_id]

# راه‌اندازی ربات
bot.polling()
